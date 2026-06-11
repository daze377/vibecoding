# Day 5 example 3 — a gradebook with live formulas (openpyxl)
# Run:  python make_grades.py   ->  grades.xlsx
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

STUDENTS = [("Amy", 9, 10, 8), ("Ben", 7, 8, 9), ("Carla", 10, 9, 10),
            ("Diego", 6, 7, 8), ("Emma", 8, 9, 9), ("Felix", 9, 9, 10),
            ("Grace", 10, 10, 9), ("Hugo", 5, 6, 7), ("Iris", 8, 8, 8),
            ("Jack", 7, 9, 8)]

wb = Workbook()
ws = wb.active
ws.title = "Grades"
ws.append(["Name", "Quiz 1", "Quiz 2", "Quiz 3", "Average"])

blue = PatternFill("solid", fgColor="0EA5E9")
green = PatternFill("solid", fgColor="C6EFCE")
for cell in ws[1]:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = blue

for i, (name, q1, q2, q3) in enumerate(STUDENTS, start=2):
    ws.append([name, q1, q2, q3])
    ws[f"E{i}"] = f"=AVERAGE(B{i}:D{i})"          # a real, live formula
    ws[f"E{i}"].number_format = "0.0"
    if (q1 + q2 + q3) / 3 > 9:                     # highlight the stars
        for col in "ABCDE":
            ws[f"{col}{i}"].fill = green

last = len(STUDENTS) + 2
ws[f"A{last}"] = "Class average"
ws[f"A{last}"].font = Font(bold=True)
for col in "BCDE":
    ws[f"{col}{last}"] = f"=AVERAGE({col}2:{col}{last-1})"
    ws[f"{col}{last}"].font = Font(bold=True)
    ws[f"{col}{last}"].number_format = "0.0"

wb.save("grades.xlsx")
print("grades.xlsx created — formulas stay live in Excel!")
